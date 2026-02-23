"""
Admin Reports Service
Builds System Admin Month + Fiscal YTD report payloads and exports.
"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.config.constants import MetricID
from src.db.models import (
    AuditLog,
    ClusterMaster,
    CompanyMaster,
    FinancialFact,
    PeriodMaster,
    ReportExportHistory,
    UserMaster,
)

logger = logging.getLogger(__name__)

_ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
_LOGO_PNG_PATH = _ASSETS_DIR / "blue-75-years-logo.png"


class ReportValidationError(Exception):
    """Raised when report filters are invalid."""


class ReportNotFoundError(Exception):
    """Raised when report data is not found."""


MONTH_NAMES = [
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

_METRIC_IDS: Dict[str, int] = {
    "revenue": int(MetricID.REVENUE),
    "gp": int(MetricID.GP),
    "other_income": int(MetricID.OTHER_INCOME),
    "personal_expenses": int(MetricID.PERSONAL_EXP),
    "admin_expenses": int(MetricID.ADMIN_EXP),
    "selling_expenses": int(MetricID.SELLING_EXP),
    "finance_expenses": int(MetricID.FINANCE_EXP),
    "depreciation": int(MetricID.DEPRECIATION),
    "provisions": int(MetricID.PROVISIONS),
    "exchange": int(MetricID.EXCHANGE_VARIANCE),
    "non_ops_expenses": int(MetricID.NON_OPS_EXP),
    "non_ops_income": int(MetricID.NON_OPS_INCOME),
}

_ROW_ORDER: List[Tuple[str, str, bool]] = [
    ("revenue", "Revenue", False),
    ("gp", "GP", False),
    ("gp_margin", "GP Margin", True),
    ("other_income", "Other income", False),
    ("personal_expenses", "Personal expenses", False),
    ("admin_expenses", "Admin expenses", False),
    ("selling_expenses", "Selling expenses", False),
    ("finance_expenses", "Finance expenses", False),
    ("depreciation", "Depreciation", False),
    ("total_overhead", "Total overhead", False),
    ("provisions", "Provisions", False),
    ("exchange", "Exchange", False),
    ("pbt_before_non_ops", "PBT before non ops", False),
    ("non_ops_expenses", "Non ops expenses", False),
    ("non_ops_income", "Non ops income", False),
    ("pbt_after_non_ops", "PBT after non ops", False),
    ("np_margin", "NP margin", True),
    ("ebit", "EBIT", False),
    ("ebitda", "EBITDA", False),
]

CONFIDENTIALITY_NOTICE = (
    "This document is issued by McLarens Group Management Pvt Limited and contains "
    "confidential financial information intended solely for authorized personnel. Any "
    "review, disclosure, copying, distribution, or use of this information by anyone other "
    "than the intended recipient is strictly prohibited. If you are not the authorized "
    "recipient, please delete this document immediately and notify the sender. Unauthorized "
    "sharing of this report may result in disciplinary or legal action."
)
COPYRIGHT_NOTICE = "(c) McLarens Group Management Pvt Limited - Business Transformation"


def _to_float(value: Any) -> float:
    return float(value) if value is not None else 0.0


def _month_name(month: int) -> str:
    if month < 1 or month > 12:
        return str(month)
    return MONTH_NAMES[month]


def _format_metric(value: float, is_percentage: bool) -> str:
    if is_percentage:
        return f"{value:,.2f}%"
    return f"{value:,.2f}"


def _achievement(actual: float, budget: float) -> float:
    if budget == 0:
        return 0.0
    return (actual / budget) * 100.0


class AdminReportService:
    """Business logic for System Admin financial reports and exports."""

    @staticmethod
    async def _get_company_and_cluster(
        db: AsyncSession,
        cluster_id: str,
        company_id: str,
    ) -> Tuple[ClusterMaster, CompanyMaster]:
        cluster = (
            await db.execute(
                select(ClusterMaster).where(
                    and_(
                        ClusterMaster.cluster_id == cluster_id,
                        ClusterMaster.is_active.is_(True),
                    )
                )
            )
        ).scalar_one_or_none()
        if not cluster:
            raise ReportNotFoundError("Cluster not found")

        company = (
            await db.execute(
                select(CompanyMaster).where(
                    and_(
                        CompanyMaster.company_id == company_id,
                        CompanyMaster.is_active.is_(True),
                    )
                )
            )
        ).scalar_one_or_none()
        if not company:
            raise ReportNotFoundError("Company not found")

        if company.cluster_id != cluster_id:
            raise ReportValidationError("Selected company does not belong to selected cluster")

        return cluster, company

    @staticmethod
    async def _get_period(db: AsyncSession, year: int, month: int) -> PeriodMaster:
        period = (
            await db.execute(
                select(PeriodMaster).where(
                    and_(PeriodMaster.year == year, PeriodMaster.month == month)
                )
            )
        ).scalar_one_or_none()
        if not period:
            raise ReportNotFoundError(f"Period not found for {year}-{month:02d}")
        return period

    @staticmethod
    async def _get_ytd_periods(
        db: AsyncSession,
        report_year: int,
        report_month: int,
        fin_year_start_month: int,
    ) -> List[PeriodMaster]:
        fy_start_month = fin_year_start_month or 1
        if report_month >= fy_start_month:
            fy_start_year = report_year
        else:
            fy_start_year = report_year - 1

        if fy_start_year == report_year:
            condition = and_(
                PeriodMaster.year == report_year,
                PeriodMaster.month >= fy_start_month,
                PeriodMaster.month <= report_month,
            )
        else:
            condition = or_(
                and_(
                    PeriodMaster.year == fy_start_year,
                    PeriodMaster.month >= fy_start_month,
                ),
                and_(
                    PeriodMaster.year == report_year,
                    PeriodMaster.month <= report_month,
                ),
            )

        result = await db.execute(
            select(PeriodMaster)
            .where(condition)
            .order_by(PeriodMaster.year.asc(), PeriodMaster.month.asc())
        )
        periods = result.scalars().all()
        if not periods:
            raise ReportNotFoundError("No periods found for fiscal YTD range")
        return periods

    @staticmethod
    async def _aggregate_metrics_by_scenario(
        db: AsyncSession,
        company_id: str,
        period_ids: List[int],
        scenario: str,
    ) -> Dict[int, float]:
        if not period_ids:
            return {}

        normalized_scenario = scenario.strip().upper()
        rows = (
            await db.execute(
                select(
                    FinancialFact.metric_id,
                    func.sum(FinancialFact.amount).label("total_amount"),
                )
                .where(
                    and_(
                        FinancialFact.company_id == company_id,
                        FinancialFact.period_id.in_(period_ids),
                        func.upper(func.trim(FinancialFact.actual_budget)) == normalized_scenario,
                    )
                )
                .group_by(FinancialFact.metric_id)
            )
        ).all()
        return {int(metric_id): _to_float(total_amount) for metric_id, total_amount in rows}

    @staticmethod
    def _compute_metric_map(metric_sums_by_id: Dict[int, float]) -> Dict[str, float]:
        base = {
            "revenue": metric_sums_by_id.get(_METRIC_IDS["revenue"], 0.0),
            "gp": metric_sums_by_id.get(_METRIC_IDS["gp"], 0.0),
            "other_income": metric_sums_by_id.get(_METRIC_IDS["other_income"], 0.0),
            "personal_expenses": metric_sums_by_id.get(_METRIC_IDS["personal_expenses"], 0.0),
            "admin_expenses": metric_sums_by_id.get(_METRIC_IDS["admin_expenses"], 0.0),
            "selling_expenses": metric_sums_by_id.get(_METRIC_IDS["selling_expenses"], 0.0),
            "finance_expenses": metric_sums_by_id.get(_METRIC_IDS["finance_expenses"], 0.0),
            "depreciation": metric_sums_by_id.get(_METRIC_IDS["depreciation"], 0.0),
            "provisions": metric_sums_by_id.get(_METRIC_IDS["provisions"], 0.0),
            "exchange": metric_sums_by_id.get(_METRIC_IDS["exchange"], 0.0),
            "non_ops_expenses": metric_sums_by_id.get(_METRIC_IDS["non_ops_expenses"], 0.0),
            "non_ops_income": metric_sums_by_id.get(_METRIC_IDS["non_ops_income"], 0.0),
        }

        total_overhead = (
            base["personal_expenses"]
            + base["admin_expenses"]
            + base["selling_expenses"]
            + base["finance_expenses"]
            + base["depreciation"]
        )
        pbt_before_non_ops = (
            base["gp"] + base["other_income"] - total_overhead + base["provisions"] + base["exchange"]
        )
        pbt_after_non_ops = pbt_before_non_ops - base["non_ops_expenses"] + base["non_ops_income"]

        revenue = base["revenue"]
        gp_margin = (base["gp"] / revenue * 100) if revenue != 0 else 0.0
        np_margin = (pbt_before_non_ops / revenue * 100) if revenue != 0 else 0.0
        ebit = pbt_before_non_ops + base["finance_expenses"]
        ebitda = ebit + base["depreciation"]

        return {
            **base,
            "gp_margin": gp_margin,
            "total_overhead": total_overhead,
            "pbt_before_non_ops": pbt_before_non_ops,
            "pbt_after_non_ops": pbt_after_non_ops,
            "np_margin": np_margin,
            "ebit": ebit,
            "ebitda": ebitda,
        }

    @staticmethod
    def _build_rows(
        month_actual_metrics: Dict[str, float],
        month_budget_metrics: Dict[str, float],
        ytd_actual_metrics: Dict[str, float],
        ytd_budget_metrics: Dict[str, float],
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for metric_key, metric_label, is_percentage in _ROW_ORDER:
            month_actual = _to_float(month_actual_metrics.get(metric_key))
            month_budget = _to_float(month_budget_metrics.get(metric_key))
            ytd_actual = _to_float(ytd_actual_metrics.get(metric_key))
            ytd_budget = _to_float(ytd_budget_metrics.get(metric_key))
            rows.append(
                {
                    "metric_key": metric_key,
                    "metric_label": metric_label,
                    "month_actual": month_actual,
                    "month_budget": month_budget,
                    "month_achievement": _achievement(month_actual, month_budget),
                    "ytd_actual": ytd_actual,
                    "ytd_budget": ytd_budget,
                    "ytd_achievement": _achievement(ytd_actual, ytd_budget),
                    "is_percentage": is_percentage,
                }
            )
        return rows

    @staticmethod
    async def get_available_years(
        db: AsyncSession,
        company_id: str,
    ) -> List[int]:
        company = (
            await db.execute(
                select(CompanyMaster).where(
                    and_(
                        CompanyMaster.company_id == company_id,
                        CompanyMaster.is_active.is_(True),
                    )
                )
            )
        ).scalar_one_or_none()
        if not company:
            raise ReportNotFoundError("Company not found")

        rows = (
            await db.execute(
                select(PeriodMaster.year)
                .join(FinancialFact, FinancialFact.period_id == PeriodMaster.period_id)
                .where(
                    and_(
                        FinancialFact.company_id == company_id,
                        func.upper(func.trim(FinancialFact.actual_budget)) == "ACTUAL",
                    )
                )
                .distinct()
                .order_by(PeriodMaster.year.desc())
            )
        ).all()
        return [int(year) for (year,) in rows]

    @staticmethod
    async def get_available_months(
        db: AsyncSession,
        company_id: str,
        year: int,
    ) -> List[Dict[str, Any]]:
        company = (
            await db.execute(
                select(CompanyMaster).where(
                    and_(
                        CompanyMaster.company_id == company_id,
                        CompanyMaster.is_active.is_(True),
                    )
                )
            )
        ).scalar_one_or_none()
        if not company:
            raise ReportNotFoundError("Company not found")

        rows = (
            await db.execute(
                select(PeriodMaster.month)
                .join(FinancialFact, FinancialFact.period_id == PeriodMaster.period_id)
                .where(
                    and_(
                        FinancialFact.company_id == company_id,
                        PeriodMaster.year == year,
                        func.upper(func.trim(FinancialFact.actual_budget)) == "ACTUAL",
                    )
                )
                .distinct()
                .order_by(PeriodMaster.month.asc())
            )
        ).all()

        return [
            {"month": int(month), "month_name": _month_name(int(month))}
            for (month,) in rows
        ]

    @staticmethod
    async def build_report_preview(
        db: AsyncSession,
        cluster_id: str,
        company_id: str,
        year: int,
        month: int,
    ) -> Dict[str, Any]:
        cluster, company = await AdminReportService._get_company_and_cluster(
            db, cluster_id, company_id
        )
        selected_period = await AdminReportService._get_period(db, year, month)

        month_actual_sums = await AdminReportService._aggregate_metrics_by_scenario(
            db=db,
            company_id=company_id,
            period_ids=[selected_period.period_id],
            scenario="ACTUAL",
        )
        if not month_actual_sums:
            raise ReportNotFoundError(
                f"No ACTUAL data found for {company.company_name} in {year}-{month:02d}"
            )
        month_budget_sums = await AdminReportService._aggregate_metrics_by_scenario(
            db=db,
            company_id=company_id,
            period_ids=[selected_period.period_id],
            scenario="BUDGET",
        )

        fin_year_start_month = int(company.fin_year_start_month or 1)
        ytd_periods = await AdminReportService._get_ytd_periods(
            db=db,
            report_year=year,
            report_month=month,
            fin_year_start_month=fin_year_start_month,
        )
        ytd_period_ids = [p.period_id for p in ytd_periods]
        ytd_actual_sums = await AdminReportService._aggregate_metrics_by_scenario(
            db=db,
            company_id=company_id,
            period_ids=ytd_period_ids,
            scenario="ACTUAL",
        )
        ytd_budget_sums = await AdminReportService._aggregate_metrics_by_scenario(
            db=db,
            company_id=company_id,
            period_ids=ytd_period_ids,
            scenario="BUDGET",
        )

        month_actual_metrics = AdminReportService._compute_metric_map(month_actual_sums)
        month_budget_metrics = AdminReportService._compute_metric_map(month_budget_sums)
        ytd_actual_metrics = AdminReportService._compute_metric_map(ytd_actual_sums)
        ytd_budget_metrics = AdminReportService._compute_metric_map(ytd_budget_sums)
        rows = AdminReportService._build_rows(
            month_actual_metrics,
            month_budget_metrics,
            ytd_actual_metrics,
            ytd_budget_metrics,
        )

        month_label = datetime(year, month, 1).strftime("%b %Y")
        ytd_start = datetime(ytd_periods[0].year, ytd_periods[0].month, 1).strftime("%b %Y")
        ytd_end = datetime(ytd_periods[-1].year, ytd_periods[-1].month, 1).strftime("%b %Y")
        period_label = datetime(year, month, 1).strftime("%B %Y")

        return {
            "cluster_id": cluster.cluster_id,
            "cluster_name": cluster.cluster_name,
            "company_id": company.company_id,
            "company_name": company.company_name,
            "year": year,
            "month": month,
            "period_label": period_label,
            "fin_year_start_month": fin_year_start_month,
            "matrix": {
                "month_label": month_label,
                "ytd_label": f"{ytd_start} - {ytd_end}",
            },
            "rows": rows,
            "month_actual_values": month_actual_metrics,
            "month_budget_values": month_budget_metrics,
            "ytd_actual_values": ytd_actual_metrics,
            "ytd_budget_values": ytd_budget_metrics,
        }

    @staticmethod
    def _add_page_footer(canvas, doc):
        """Draw footer on every page: timestamp left, page number right."""
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748b"))
        timestamp = datetime.utcnow().strftime("Generated: %d %b %Y, %H:%M UTC")
        canvas.drawString(18 * mm, 10 * mm, timestamp)
        page_text = f"Page {canvas.getPageNumber()}"
        canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, page_text)
        canvas.restoreState()

    @staticmethod
    def generate_pdf(preview: Dict[str, Any]) -> bytes:
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )

        styles = getSampleStyleSheet()
        story = []

        # --- Logo ---
        try:
            if _LOGO_PNG_PATH.exists():
                logo = PdfImage(str(_LOGO_PNG_PATH), width=50 * mm, height=25 * mm)
                story.append(logo)
                story.append(Spacer(1, 6))
        except Exception:
            logger.warning("Could not load PNG logo for PDF; proceeding without it")

        story.append(Paragraph("McLarens Analytics - Financial Report", styles["Title"]))
        story.append(Spacer(1, 8))

        metadata_table = Table(
            [
                ["Cluster", preview["cluster_name"], "Company", preview["company_name"]],
                ["Period", preview["period_label"], "Matrix", "Month + YTD"],
            ],
            colWidths=[24 * mm, 63 * mm, 24 * mm, 63 * mm],
        )
        metadata_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#eef2ff")),
                    ("BACKGROUND", (0, 1), (3, 1), colors.HexColor("#f8fafc")),
                    ("FONTNAME", (0, 0), (3, 1), "Helvetica"),
                    ("FONTNAME", (0, 0), (0, 1), "Helvetica-Bold"),
                    ("FONTNAME", (2, 0), (2, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (3, 1), 9),
                    ("GRID", (0, 0), (3, 1), 0.4, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (3, 1), "MIDDLE"),
                ]
            )
        )
        story.append(metadata_table)
        story.append(Spacer(1, 10))

        rows = [
            ["Metric", preview["matrix"]["month_label"], "", "", "YTD", "", ""],
            ["", "Actual", "Budget", "Achievement", "Actual", "Budget", "Achievement"],
        ]
        for row in preview["rows"]:
            rows.append(
                [
                    row["metric_label"],
                    _format_metric(_to_float(row["month_actual"]), bool(row["is_percentage"])),
                    _format_metric(_to_float(row["month_budget"]), bool(row["is_percentage"])),
                    _format_metric(_to_float(row["month_achievement"]), True),
                    _format_metric(_to_float(row["ytd_actual"]), bool(row["is_percentage"])),
                    _format_metric(_to_float(row["ytd_budget"]), bool(row["is_percentage"])),
                    _format_metric(_to_float(row["ytd_achievement"]), True),
                ]
            )

        report_table = Table(
            rows,
            colWidths=[45 * mm, 21.5 * mm, 21.5 * mm, 21.5 * mm, 21.5 * mm, 21.5 * mm, 21.5 * mm],
        )
        report_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (6, 0), colors.HexColor("#0b1f3a")),
                    ("BACKGROUND", (0, 1), (6, 1), colors.HexColor("#1e3a5f")),
                    ("TEXTCOLOR", (0, 0), (6, 1), colors.white),
                    ("FONTNAME", (0, 0), (6, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (6, 1), 8.5),
                    ("SPAN", (1, 0), (3, 0)),
                    ("SPAN", (4, 0), (6, 0)),
                    ("SPAN", (0, 0), (0, 1)),
                    ("VALIGN", (0, 0), (6, 1), "MIDDLE"),
                    ("ALIGN", (1, 0), (6, -1), "RIGHT"),
                    ("ALIGN", (1, 0), (6, 1), "CENTER"),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("FONTNAME", (0, 2), (6, -1), "Helvetica"),
                    ("FONTSIZE", (0, 2), (6, -1), 8),
                    ("ROWBACKGROUNDS", (0, 2), (6, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0, 0), (6, -1), 0.35, colors.HexColor("#cbd5e1")),
                    ("TOPPADDING", (0, 0), (6, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (6, -1), 4),
                ]
            )
        )
        story.append(report_table)
        story.append(Spacer(1, 8))

        footer_style = ParagraphStyle(
            "ConfidentialFooter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor("#475569"),
        )
        story.append(Paragraph(CONFIDENTIALITY_NOTICE, footer_style))
        story.append(Spacer(1, 2))
        story.append(Paragraph(COPYRIGHT_NOTICE, footer_style))

        document.build(
            story,
            onFirstPage=AdminReportService._add_page_footer,
            onLaterPages=AdminReportService._add_page_footer,
        )
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel(preview: Dict[str, Any]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Financial Report"

        title_font = Font(bold=True, size=14, color="0B1F3A")
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", start_color="0B1F3A", end_color="0B1F3A")
        sub_header_fill = PatternFill(fill_type="solid", start_color="1E3A5F", end_color="1E3A5F")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --- Logo ---
        logo_row_offset = 0
        try:
            if _LOGO_PNG_PATH.exists():
                img = XlImage(str(_LOGO_PNG_PATH))
                img.width = 150
                img.height = 75
                sheet.add_image(img, "A1")
                logo_row_offset = 5
        except Exception:
            logger.warning("Could not load PNG logo for Excel; proceeding without it")

        title_row = 1 + logo_row_offset
        sheet.merge_cells(
            start_row=title_row, start_column=1,
            end_row=title_row, end_column=7,
        )
        title_cell = sheet.cell(row=title_row, column=1, value="McLarens Analytics - Financial Report")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center")

        meta_start = title_row + 2
        sheet.cell(row=meta_start, column=1, value="Cluster").font = bold_font
        sheet.cell(row=meta_start, column=2, value=preview["cluster_name"])
        sheet.cell(row=meta_start, column=4, value="Company").font = bold_font
        sheet.cell(row=meta_start, column=5, value=preview["company_name"])
        sheet.cell(row=meta_start + 1, column=1, value="Period").font = bold_font
        sheet.cell(row=meta_start + 1, column=2, value=preview["period_label"])
        sheet.cell(row=meta_start + 1, column=4, value="Matrix").font = bold_font
        sheet.cell(row=meta_start + 1, column=5, value="Month + YTD")

        header_row = meta_start + 3
        sub_header_row = header_row + 1
        sheet.cell(row=header_row, column=1, value="Metric")
        sheet.cell(row=header_row, column=2, value=preview["matrix"]["month_label"])
        sheet.cell(row=header_row, column=5, value=preview["matrix"]["ytd_label"])
        sheet.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=4)
        sheet.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=7)
        sheet.merge_cells(start_row=header_row, start_column=1, end_row=sub_header_row, end_column=1)

        sub_headers = ["Actual", "Budget", "Achievement", "Actual", "Budget", "Achievement"]
        for index, label in enumerate(sub_headers, start=2):
            sheet.cell(row=sub_header_row, column=index, value=label)

        for col in range(1, 8):
            cell = sheet.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for col in range(2, 8):
            cell = sheet.cell(row=sub_header_row, column=col)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        metric_header_cell = sheet.cell(row=sub_header_row, column=1)
        metric_header_cell.fill = header_fill
        metric_header_cell.border = border

        current_row = sub_header_row + 1
        for row in preview["rows"]:
            sheet.cell(row=current_row, column=1, value=row["metric_label"]).border = border

            month_actual_cell = sheet.cell(row=current_row, column=2, value=_to_float(row["month_actual"]))
            month_budget_cell = sheet.cell(row=current_row, column=3, value=_to_float(row["month_budget"]))
            month_achievement_cell = sheet.cell(row=current_row, column=4, value=_to_float(row["month_achievement"]))
            ytd_actual_cell = sheet.cell(row=current_row, column=5, value=_to_float(row["ytd_actual"]))
            ytd_budget_cell = sheet.cell(row=current_row, column=6, value=_to_float(row["ytd_budget"]))
            ytd_achievement_cell = sheet.cell(row=current_row, column=7, value=_to_float(row["ytd_achievement"]))

            data_cells = [
                month_actual_cell,
                month_budget_cell,
                month_achievement_cell,
                ytd_actual_cell,
                ytd_budget_cell,
                ytd_achievement_cell,
            ]
            for cell in data_cells:
                cell.border = border
                cell.alignment = Alignment(horizontal="right")

            if row["is_percentage"]:
                month_actual_cell.number_format = "0.00%"
                month_budget_cell.number_format = "0.00%"
                month_actual_cell.value = _to_float(row["month_actual"]) / 100.0
                month_budget_cell.value = _to_float(row["month_budget"]) / 100.0
                ytd_actual_cell.number_format = "0.00%"
                ytd_budget_cell.number_format = "0.00%"
                ytd_actual_cell.value = _to_float(row["ytd_actual"]) / 100.0
                ytd_budget_cell.value = _to_float(row["ytd_budget"]) / 100.0
            else:
                month_actual_cell.number_format = "#,##0.00"
                month_budget_cell.number_format = "#,##0.00"
                ytd_actual_cell.number_format = "#,##0.00"
                ytd_budget_cell.number_format = "#,##0.00"

            month_achievement_cell.number_format = "0.00%"
            month_achievement_cell.value = _to_float(row["month_achievement"]) / 100.0
            ytd_achievement_cell.number_format = "0.00%"
            ytd_achievement_cell.value = _to_float(row["ytd_achievement"]) / 100.0
            current_row += 1

        sheet.column_dimensions["A"].width = 36
        for col in ("B", "C", "D", "E", "F", "G"):
            sheet.column_dimensions[col].width = 15

        footer_row = current_row + 2
        sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row + 2, end_column=7)
        footer_cell = sheet.cell(row=footer_row, column=1, value=f"{CONFIDENTIALITY_NOTICE}\n\n{COPYRIGHT_NOTICE}")
        footer_cell.alignment = Alignment(wrap_text=True, vertical="top")
        footer_cell.font = Font(size=9, color="475569")

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    async def record_export(
        db: AsyncSession,
        *,
        exported_by: str,
        export_format: str,
        file_name: str,
        preview: Dict[str, Any],
    ) -> ReportExportHistory:
        history = ReportExportHistory(
            exported_by=exported_by,
            export_format=export_format,
            file_name=file_name,
            cluster_id=preview["cluster_id"],
            cluster_name=preview["cluster_name"],
            company_id=preview["company_id"],
            company_name=preview["company_name"],
            year=preview["year"],
            month=preview["month"],
            period_label=preview["period_label"],
        )
        db.add(history)

        action = "REPORT_EXPORT_PDF" if export_format.upper() == "PDF" else "REPORT_EXPORT_EXCEL"
        db.add(
            AuditLog(
                user_id=exported_by,
                action=action,
                entity_type="report_export",
                entity_id=f"{preview['company_id']}:{preview['year']}-{preview['month']:02d}",
                details=f"{export_format.upper()} export for {preview['company_name']} ({preview['period_label']})",
            )
        )

        await db.flush()
        return history

    @staticmethod
    async def list_export_history(
        db: AsyncSession,
        *,
        limit: int,
        offset: int,
    ) -> Dict[str, Any]:
        total = (
            await db.execute(select(func.count(ReportExportHistory.id)))
        ).scalar() or 0

        rows = (
            await db.execute(
                select(
                    ReportExportHistory,
                    UserMaster.user_email,
                    UserMaster.first_name,
                    UserMaster.last_name,
                )
                .outerjoin(UserMaster, UserMaster.user_id == ReportExportHistory.exported_by)
                .order_by(ReportExportHistory.exported_at.desc())
                .offset(offset)
                .limit(limit)
            )
        ).all()

        history_items = []
        for history, user_email, first_name, last_name in rows:
            display_name = None
            if first_name or last_name:
                display_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
            if not display_name:
                display_name = user_email or history.exported_by

            history_items.append(
                {
                    "id": history.id,
                    "exported_at": history.exported_at,
                    "exported_by": history.exported_by,
                    "exported_by_email": user_email,
                    "exported_by_name": display_name,
                    "export_format": history.export_format,
                    "file_name": history.file_name,
                    "cluster_id": history.cluster_id,
                    "cluster_name": history.cluster_name,
                    "company_id": history.company_id,
                    "company_name": history.company_name,
                    "year": history.year,
                    "month": history.month,
                    "period_label": history.period_label,
                }
            )

        return {
            "items": history_items,
            "total": int(total),
            "limit": limit,
            "offset": offset,
        }
