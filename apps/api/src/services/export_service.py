"""
Export Service for McLarens Analytics
Generates Excel exports from database data
"""

import io
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from db.models import Cluster, Company, FinancialData


class ExportService:
    """Service for generating various export formats"""
    
    # Template path
    TEMPLATE_DIR = Path(__file__).parent.parent.parent / "seed data csv"
    
    @staticmethod
    async def get_group_financial_summary(
        db: AsyncSession,
        year: int,
        month: int
    ) -> Dict:
        """
        Get aggregated financial data for Group Financial Summary export.
        Returns data organized by cluster.
        """
        # Get all clusters with companies and financial data
        result = await db.execute(
            select(Cluster)
            .options(selectinload(Cluster.companies))
            .where(Cluster.is_active == True)
        )
        clusters = result.scalars().all()
        
        summary = {
            "period": f"{datetime(year, month, 1).strftime('%B')} {year}",
            "year": year,
            "month": month,
            "clusters": [],
            "totals": {
                "revenue_actual": 0,
                "revenue_budget": 0,
                "gp_actual": 0,
                "gp_budget": 0,
                "pbt_actual": 0,
                "pbt_budget": 0,
                "ebitda_actual": 0,
                "ebitda_budget": 0,
            }
        }
        
        for cluster in clusters:
            cluster_data = {
                "name": cluster.name,
                "code": cluster.code,
                "companies": [],
                "totals": {
                    "revenue_actual": 0,
                    "revenue_budget": 0,
                    "gp_actual": 0,
                    "gp_budget": 0,
                    "pbt_actual": 0,
                    "pbt_budget": 0,
                    "ebitda_actual": 0,
                    "ebitda_budget": 0,
                }
            }
            
            for company in cluster.companies:
                if not company.is_active:
                    continue
                
                # Get financial data for this company/period
                fin_result = await db.execute(
                    select(FinancialData)
                    .where(
                        FinancialData.company_id == company.id,
                        FinancialData.year == year,
                        FinancialData.month == month
                    )
                )
                fin_data = fin_result.scalar_one_or_none()
                
                company_data = {
                    "code": company.code,
                    "name": company.name,
                    "revenue_actual": fin_data.revenue_lkr_actual if fin_data else 0,
                    "revenue_budget": fin_data.revenue_lkr_budget if fin_data else 0,
                    "gp_actual": fin_data.gp_actual if fin_data else 0,
                    "gp_budget": fin_data.gp_budget if fin_data else 0,
                    "pbt_actual": fin_data.pbt_before_actual if fin_data else 0,
                    "pbt_budget": fin_data.pbt_before_budget if fin_data else 0,
                    "ebitda_actual": fin_data.ebitda_computed_actual if fin_data else 0,
                    "ebitda_budget": fin_data.ebitda_computed_budget if fin_data else 0,
                }
                
                cluster_data["companies"].append(company_data)
                
                # Add to cluster totals
                for key in cluster_data["totals"]:
                    cluster_data["totals"][key] += company_data[key]
            
            summary["clusters"].append(cluster_data)
            
            # Add to group totals
            for key in summary["totals"]:
                summary["totals"][key] += cluster_data["totals"][key]
        
        return summary
    
    @staticmethod
    async def generate_excel_report(
        db: AsyncSession,
        year: int,
        month: int
    ) -> bytes:
        """
        Generate Excel workbook with Group Financial Summary.
        Uses the P&L Template format.
        """
        # Get data
        summary = await ExportService.get_group_financial_summary(db, year, month)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Group P&L Summary"
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        cluster_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"GROUP FINANCIAL SUMMARY - {summary['period']}"
        ws['A1'].font = Font(bold=True, size=16, color="0B1F3A")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Generated timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10, color="64748B")
        
        # Headers
        headers = [
            "Cluster / Company",
            "Revenue (Actual)",
            "Revenue (Budget)",
            "GP (Actual)",
            "GP (Budget)",
            "PBT (Actual)",
            "PBT (Budget)",
            "Variance %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        # Data rows
        row_num = 5
        
        for cluster in summary["clusters"]:
            # Cluster header row
            ws.cell(row=row_num, column=1).value = cluster["name"]
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row_num, column=1).fill = cluster_fill
            
            ws.cell(row=row_num, column=2).value = cluster["totals"]["revenue_actual"]
            ws.cell(row=row_num, column=3).value = cluster["totals"]["revenue_budget"]
            ws.cell(row=row_num, column=4).value = cluster["totals"]["gp_actual"]
            ws.cell(row=row_num, column=5).value = cluster["totals"]["gp_budget"]
            ws.cell(row=row_num, column=6).value = cluster["totals"]["pbt_actual"]
            ws.cell(row=row_num, column=7).value = cluster["totals"]["pbt_budget"]
            
            # Calculate variance
            if cluster["totals"]["pbt_budget"] != 0:
                variance = ((cluster["totals"]["pbt_actual"] - cluster["totals"]["pbt_budget"]) 
                           / abs(cluster["totals"]["pbt_budget"]) * 100)
                ws.cell(row=row_num, column=8).value = f"{variance:.1f}%"
            else:
                ws.cell(row=row_num, column=8).value = "-"
            
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = cluster_fill
                ws.cell(row=row_num, column=col).border = border
                if col > 1:
                    ws.cell(row=row_num, column=col).number_format = '#,##0'
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')
            
            row_num += 1
            
            # Company rows
            for company in cluster["companies"]:
                ws.cell(row=row_num, column=1).value = f"  {company['name']}"
                ws.cell(row=row_num, column=2).value = company["revenue_actual"]
                ws.cell(row=row_num, column=3).value = company["revenue_budget"]
                ws.cell(row=row_num, column=4).value = company["gp_actual"]
                ws.cell(row=row_num, column=5).value = company["gp_budget"]
                ws.cell(row=row_num, column=6).value = company["pbt_actual"]
                ws.cell(row=row_num, column=7).value = company["pbt_budget"]
                
                if company["pbt_budget"] != 0:
                    variance = ((company["pbt_actual"] - company["pbt_budget"]) 
                               / abs(company["pbt_budget"]) * 100)
                    ws.cell(row=row_num, column=8).value = f"{variance:.1f}%"
                else:
                    ws.cell(row=row_num, column=8).value = "-"
                
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).border = border
                    if col > 1:
                        ws.cell(row=row_num, column=col).number_format = '#,##0'
                        ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')
                
                row_num += 1
        
        # Group totals row
        row_num += 1
        ws.cell(row=row_num, column=1).value = "GROUP TOTAL"
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=row_num, column=2).value = summary["totals"]["revenue_actual"]
        ws.cell(row=row_num, column=3).value = summary["totals"]["revenue_budget"]
        ws.cell(row=row_num, column=4).value = summary["totals"]["gp_actual"]
        ws.cell(row=row_num, column=5).value = summary["totals"]["gp_budget"]
        ws.cell(row=row_num, column=6).value = summary["totals"]["pbt_actual"]
        ws.cell(row=row_num, column=7).value = summary["totals"]["pbt_budget"]
        
        if summary["totals"]["pbt_budget"] != 0:
            variance = ((summary["totals"]["pbt_actual"] - summary["totals"]["pbt_budget"]) 
                       / abs(summary["totals"]["pbt_budget"]) * 100)
            ws.cell(row=row_num, column=8).value = f"{variance:.1f}%"
        
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = subheader_fill
            cell.border = border
            if col > 1:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    async def get_available_periods(db: AsyncSession) -> List[Dict]:
        """Get list of periods with financial data available"""
        result = await db.execute(
            select(
                FinancialData.year,
                FinancialData.month,
                func.count(FinancialData.id).label('count')
            )
            .group_by(FinancialData.year, FinancialData.month)
            .order_by(FinancialData.year.desc(), FinancialData.month.desc())
        )
        
        periods = []
        for row in result.all():
            month_name = datetime(row[0], row[1], 1).strftime('%B')
            periods.append({
                "year": row[0],
                "month": row[1],
                "period": f"{month_name} {row[0]}",
                "company_count": row[2]
            })
        
        return periods
