"""
Data Seeding Script for McLarens Analytics
This script imports data from CSV/Excel files into the database.

Files processed:
- Center_CompanyMaster.csv: Companies and Clusters
- Users.xlsx: Users with their company/cluster assignments
- Monthly Data Files (April-October 2025): Financial data

Usage:
    python seed_real_data.py
"""

import asyncio
import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import uuid
from decimal import Decimal

# Third-party imports
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

# Local imports - adjust path as needed
import sys
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from db.models import Base, Cluster, Company, User, FinancialData, Report, UserRole, FiscalCycle, ReportStatus

# ============ CONFIGURATION ============

DATA_DIR = Path(__file__).parent / "seed data csv"
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+asyncpg://postgres:postgres@localhost:5432/mclarens_db")

# Map cluster names from CSV to normalized names
CLUSTER_NAME_MAP = {
    "Liner": "Liner",
    "Shipping services & Logistics ": "Shipping Services",
    "Shipping services & Logistics": "Shipping Services", 
    "GAC Cluster": "GAC Group",
    "Warehouse and Logistics": "Warehouse",
    "Ship Supply Services": "Ship Supply",
    "Lubricant II": "Lube 02",
    "Lubricant I": "Lube 01",
    "Manufacturing ": "Manufacturing",
    "Manufacturing": "Manufacturing",
    "Bunkering & Renewables ": "Bunkering",
    "Bunkering & Renewables": "Bunkering",
    "Property": "Property",
    "Hotel & Leisure": "Hotel & Leisure",
    "Strategic Investment": "Strategic Investment",
    "MHL and related companies": "MHL Group",
    "Dormant Companies ": "Dormant",
    "Dormant Companies": "Dormant",
}

# Map fiscal year end months
FISCAL_CYCLE_MAP = {
    "December": FiscalCycle.DECEMBER,
    "March": FiscalCycle.MARCH,
    "": FiscalCycle.MARCH,  # Default
    None: FiscalCycle.MARCH,
}


# ============ DATABASE CONNECTION ============

async def get_engine():
    """Create async database engine"""
    engine = create_async_engine(DATABASE_URL, echo=False)
    return engine


async def init_db(engine):
    """Initialize database tables"""
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    print("✅ Database tables initialized")


# ============ SEED CLUSTERS ============

async def seed_clusters(session: AsyncSession) -> Dict[str, uuid.UUID]:
    """
    Seed clusters from unique cluster names in the company master.
    Returns a mapping of cluster name -> cluster ID
    """
    print("\n📦 Seeding Clusters...")
    
    cluster_ids: Dict[str, uuid.UUID] = {}
    
    # Read unique clusters from CSV
    csv_path = DATA_DIR / "Center_CompanyMaster.csv"
    clusters_found = set()
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            cluster_raw = row.get('Cluster', '').strip()
            if cluster_raw:
                cluster_name = CLUSTER_NAME_MAP.get(cluster_raw, cluster_raw)
                clusters_found.add(cluster_name)
    
    # Check existing clusters
    result = await session.execute(select(Cluster))
    existing_clusters = {c.name: c.id for c in result.scalars().all()}
    
    # Create new clusters
    for cluster_name in sorted(clusters_found):
        if cluster_name in existing_clusters:
            cluster_ids[cluster_name] = existing_clusters[cluster_name]
            print(f"   ✓ Cluster exists: {cluster_name}")
        else:
            # Generate a code from the name
            code = cluster_name.upper().replace(' ', '_').replace('&', 'AND')[:20]
            cluster = Cluster(
                id=uuid.uuid4(),
                name=cluster_name,
                code=code,
                fiscal_cycle=FiscalCycle.DECEMBER,
                is_active=True
            )
            session.add(cluster)
            cluster_ids[cluster_name] = cluster.id
            print(f"   + Created cluster: {cluster_name}")
    
    await session.commit()
    print(f"   Total clusters: {len(cluster_ids)}")
    return cluster_ids


# ============ SEED COMPANIES ============

async def seed_companies(session: AsyncSession, cluster_ids: Dict[str, uuid.UUID]) -> Dict[str, uuid.UUID]:
    """
    Seed companies from Center_CompanyMaster.csv
    Returns a mapping of company code -> company ID
    """
    print("\n🏢 Seeding Companies...")
    
    company_ids: Dict[str, uuid.UUID] = {}
    csv_path = DATA_DIR / "Center_CompanyMaster.csv"
    
    # Check existing companies
    result = await session.execute(select(Company))
    existing_companies = {c.code: c.id for c in result.scalars().all()}
    
    created_count = 0
    skipped_count = 0
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get('CompanyCode', '').strip()
            name = row.get('CompanyName', '').strip()
            cluster_raw = row.get('Cluster', '').strip()
            is_active = row.get('IsActive', 'Yes').strip().lower() == 'yes'
            
            # Skip rows without code or name
            if not code or not name:
                skipped_count += 1
                continue
            
            # Get normalized cluster name
            cluster_name = CLUSTER_NAME_MAP.get(cluster_raw, cluster_raw)
            cluster_id = cluster_ids.get(cluster_name)
            
            if not cluster_id:
                print(f"   ⚠ Unknown cluster '{cluster_raw}' for company {code}, skipping")
                skipped_count += 1
                continue
            
            if code in existing_companies:
                company_ids[code] = existing_companies[code]
            else:
                company = Company(
                    id=uuid.uuid4(),
                    code=code,
                    name=name,
                    cluster_id=cluster_id,
                    is_active=is_active
                )
                session.add(company)
                company_ids[code] = company.id
                created_count += 1
    
    await session.commit()
    print(f"   Created: {created_count}, Existing: {len(existing_companies)}, Skipped: {skipped_count}")
    return company_ids


# ============ SEED USERS ============

async def seed_users(session: AsyncSession, company_ids: Dict[str, uuid.UUID], cluster_ids: Dict[str, uuid.UUID]) -> int:
    """
    Seed users from Users.xlsx
    Creates users with appropriate roles and company/cluster assignments
    """
    print("\n👥 Seeding Users...")
    
    xlsx_path = DATA_DIR / "Users.xlsx"
    
    if not xlsx_path.exists():
        print("   ⚠ Users.xlsx not found, creating default users")
        return await create_default_users(session, company_ids, cluster_ids)
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        # Check existing users
        result = await session.execute(select(User))
        existing_emails = {u.email.lower() for u in result.scalars().all()}
        
        created_count = 0
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            
            email = str(row_data.get('Email', '') or '').strip().lower()
            name = str(row_data.get('Name', '') or '').strip()
            role_str = str(row_data.get('Role', 'data_officer') or 'data_officer').strip().lower()
            
            if not email or not name:
                continue
            
            if email in existing_emails:
                continue
            
            # Map role
            role_map = {
                'ceo': UserRole.CEO,
                'admin': UserRole.ADMIN,
                'company_director': UserRole.COMPANY_DIRECTOR,
                'data_officer': UserRole.DATA_OFFICER,
                'director': UserRole.COMPANY_DIRECTOR,
            }
            role = role_map.get(role_str, UserRole.DATA_OFFICER)
            
            # Get company/cluster assignment
            company_code = str(row_data.get('CompanyCode', '') or '').strip()
            cluster_name = str(row_data.get('Cluster', '') or '').strip()
            
            company_id = company_ids.get(company_code) if company_code else None
            cluster_id = cluster_ids.get(CLUSTER_NAME_MAP.get(cluster_name, cluster_name)) if cluster_name else None
            
            user = User(
                id=uuid.uuid4(),
                email=email,
                name=name,
                password_hash="$2b$12$DevModeHash.NotForProduction",  # Placeholder
                role=role,
                company_id=company_id,
                cluster_id=cluster_id,
                is_active=True
            )
            session.add(user)
            created_count += 1
        
        await session.commit()
        wb.close()
        print(f"   Created: {created_count} users")
        return created_count
        
    except Exception as e:
        print(f"   ⚠ Error reading Users.xlsx: {e}")
        return await create_default_users(session, company_ids, cluster_ids)


async def create_default_users(session: AsyncSession, company_ids: Dict[str, uuid.UUID], cluster_ids: Dict[str, uuid.UUID]) -> int:
    """Create default users for testing"""
    
    default_users = [
        {"email": "ceo@maclarens.com", "name": "CEO User", "role": UserRole.CEO},
        {"email": "admin@maclarens.com", "name": "Admin User", "role": UserRole.ADMIN},
        {"email": "director.liner@maclarens.com", "name": "Liner Director", "role": UserRole.COMPANY_DIRECTOR, "cluster": "Liner"},
        {"email": "director.lube01@maclarens.com", "name": "Lube 01 Director", "role": UserRole.COMPANY_DIRECTOR, "cluster": "Lube 01"},
        {"email": "do.one@maclarens.com", "name": "ONE Data Officer", "role": UserRole.DATA_OFFICER, "company": "ONE"},
    ]
    
    result = await session.execute(select(User))
    existing_emails = {u.email.lower() for u in result.scalars().all()}
    
    created = 0
    for u in default_users:
        if u["email"].lower() in existing_emails:
            continue
            
        user = User(
            id=uuid.uuid4(),
            email=u["email"],
            name=u["name"],
            password_hash="$2b$12$DevModeHash.NotForProduction",
            role=u["role"],
            company_id=company_ids.get(u.get("company")) if u.get("company") else None,
            cluster_id=cluster_ids.get(u.get("cluster")) if u.get("cluster") else None,
            is_active=True
        )
        session.add(user)
        created += 1
    
    await session.commit()
    print(f"   Created: {created} default users")
    return created


# ============ SEED FINANCIAL DATA ============

async def seed_financial_data(session: AsyncSession, company_ids: Dict[str, uuid.UUID]) -> int:
    """
    Seed financial data from monthly Excel files.
    Parses files like: "April 2025 1.xlsx", "October 2025 1.xlsx", etc.
    """
    print("\n💰 Seeding Financial Data...")
    
    # Month name to number mapping
    month_map = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8, "Aug": 8,
        "September": 9, "Sep": 9, "October": 10, "November": 11, "December": 12
    }
    
    # Find all monthly data files
    monthly_files = []
    for file in DATA_DIR.glob("*.xlsx"):
        filename = file.stem
        # Parse filename like "April 2025 1" or "Oct 2025 1"
        for month_name, month_num in month_map.items():
            if filename.startswith(month_name):
                try:
                    parts = filename.split()
                    year = int(parts[1])
                    monthly_files.append((file, year, month_num))
                    break
                except (IndexError, ValueError):
                    continue
    
    if not monthly_files:
        print("   ⚠ No monthly data files found")
        return 0
    
    total_created = 0
    
    for file_path, year, month in sorted(monthly_files, key=lambda x: (x[1], x[2])):
        print(f"   Processing: {file_path.name} ({year}-{month:02d})")
        
        try:
            count = await process_monthly_file(session, file_path, year, month, company_ids)
            total_created += count
        except Exception as e:
            print(f"      ⚠ Error processing {file_path.name}: {e}")
    
    print(f"   Total financial records created: {total_created}")
    return total_created


async def process_monthly_file(
    session: AsyncSession, 
    file_path: Path, 
    year: int, 
    month: int, 
    company_ids: Dict[str, uuid.UUID]
) -> int:
    """Process a single monthly Excel file and extract financial data"""
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    created_count = 0
    
    # Check for existing data
    existing_result = await session.execute(
        select(FinancialData).where(
            FinancialData.year == year,
            FinancialData.month == month
        )
    )
    existing_companies = {str(fd.company_id) for fd in existing_result.scalars().all()}
    
    # Try to find financial data in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Look for company codes in the sheet
        for company_code, company_id in company_ids.items():
            if str(company_id) in existing_companies:
                continue
            
            # Search for company data in sheet
            financial_data = extract_company_financials(ws, company_code)
            
            if financial_data:
                fd = FinancialData(
                    id=uuid.uuid4(),
                    company_id=company_id,
                    year=year,
                    month=month,
                    **financial_data
                )
                session.add(fd)
                created_count += 1
    
    await session.commit()
    wb.close()
    
    print(f"      → Created {created_count} records")
    return created_count


def extract_company_financials(ws, company_code: str) -> Optional[Dict]:
    """
    Extract financial data for a company from a worksheet.
    This is a simplified version - you may need to adjust based on actual Excel structure.
    """
    # Search for company code in the worksheet
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == company_code.upper():
                # Found the company, try to extract data from same row or nearby
                row_num = cell.row
                
                # Sample extraction (adjust column numbers based on actual template)
                try:
                    return {
                        'revenue_lkr_actual': float(ws.cell(row=row_num, column=3).value or 0),
                        'gp_actual': float(ws.cell(row=row_num, column=4).value or 0),
                        'personal_exp_actual': float(ws.cell(row=row_num, column=5).value or 0),
                        'admin_exp_actual': float(ws.cell(row=row_num, column=6).value or 0),
                        'selling_exp_actual': float(ws.cell(row=row_num, column=7).value or 0),
                        'finance_exp_actual': float(ws.cell(row=row_num, column=8).value or 0),
                        'depreciation_actual': float(ws.cell(row=row_num, column=9).value or 0),
                        'pbt_actual': float(ws.cell(row=row_num, column=10).value or 0),
                        # Add budget columns as needed
                        'revenue_lkr_budget': float(ws.cell(row=row_num, column=11).value or 0),
                        'gp_budget': float(ws.cell(row=row_num, column=12).value or 0),
                        'pbt_budget': float(ws.cell(row=row_num, column=13).value or 0),
                    }
                except (ValueError, TypeError):
                    pass
    
    return None


# ============ CREATE SAMPLE REPORTS ============

async def seed_reports(session: AsyncSession, company_ids: Dict[str, uuid.UUID]) -> int:
    """Create sample reports for companies"""
    print("\n📝 Seeding Reports...")
    
    # Check existing reports
    result = await session.execute(select(Report))
    existing = {(str(r.company_id), r.year, r.month) for r in result.scalars().all()}
    
    created = 0
    months = [(2025, 4), (2025, 5), (2025, 6), (2025, 8), (2025, 9), (2025, 10)]
    
    for company_code, company_id in list(company_ids.items())[:20]:  # Limit to 20 companies
        for year, month in months:
            key = (str(company_id), year, month)
            if key in existing:
                continue
            
            # Random status - older months are approved
            status = ReportStatus.APPROVED if month < 10 else ReportStatus.SUBMITTED if month < 11 else ReportStatus.DRAFT
            
            report = Report(
                id=uuid.uuid4(),
                company_id=company_id,
                year=year,
                month=month,
                status=status,
                submitted_at=datetime.now() if status != ReportStatus.DRAFT else None,
                approved_at=datetime.now() if status == ReportStatus.APPROVED else None,
            )
            session.add(report)
            created += 1
    
    await session.commit()
    print(f"   Created: {created} reports")
    return created


# ============ MAIN SEED FUNCTION ============

async def main():
    """Main seeding function"""
    print("=" * 60)
    print("🌱 McLarens Analytics - Data Seeding Script")
    print("=" * 60)
    print(f"Database: {DATABASE_URL}")
    print(f"Data Directory: {DATA_DIR}")
    print("=" * 60)
    
    # Verify data directory exists
    if not DATA_DIR.exists():
        print(f"❌ Data directory not found: {DATA_DIR}")
        return
    
    # List available files
    print("\n📁 Available data files:")
    for f in DATA_DIR.glob("*.*"):
        print(f"   - {f.name}")
    
    # Create database engine
    engine = await get_engine()
    
    # Initialize database
    await init_db(engine)
    
    # Create session
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    
    async with async_session() as session:
        # Seed in order of dependencies
        cluster_ids = await seed_clusters(session)
        company_ids = await seed_companies(session, cluster_ids)
        await seed_users(session, company_ids, cluster_ids)
        await seed_financial_data(session, company_ids)
        await seed_reports(session, company_ids)
    
    await engine.dispose()
    
    print("\n" + "=" * 60)
    print("✅ Data seeding complete!")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
